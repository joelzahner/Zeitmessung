import customtkinter as ctk
import pandas as pd
import os
from datetime import datetime, timedelta
from tkinter import messagebox
from openpyxl.worksheet.header_footer import HeaderFooter
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment, Border

TABELLEN_ORDNER = "Datenbank"
AUSGABE_DATEI = os.path.join(TABELLEN_ORDNER, "Rangliste.xlsx")
DISTANZ_M = 8480  # Streckenlänge in Metern

KATEGORIEN_REIHENFOLGE = [
    "Herren",
    "Damen",
    "Junioren",
    "Juniorinnen",
    "Nachwuchs Knaben",
    "Nachwuchs Mädchen",
    "Gäste Herren",
    "Gäste Damen",
]

class RanglisteFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self.label = ctk.CTkLabel(self, text="Ranglistenerstellung", font=("Arial", 20))
        self.label.grid(row=0, column=0, pady=10)

        self.generieren_button = ctk.CTkButton(self, text="Rangliste generieren", command=self.generiere_rangliste)
        self.generieren_button.grid(row=1, column=0, pady=10)

        self.anzeige = ctk.CTkTextbox(self)
        self.anzeige.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)

    def generiere_rangliste(self):
        def lade_race(prefix):
            start_files = [f for f in os.listdir(TABELLEN_ORDNER) if f.startswith(f"Zeitmessung_Start_{prefix}")]
            ziel_files = [f for f in os.listdir(TABELLEN_ORDNER) if f.startswith(f"Zeitmessung_Ziel_{prefix}")]
            if not start_files or not ziel_files:
                return None
            start_df = pd.read_csv(
                os.path.join(TABELLEN_ORDNER, start_files[-1]),
                sep=";",
                encoding="utf-8-sig",
                dtype=str,
            )
            ziel_df = pd.read_csv(
                os.path.join(TABELLEN_ORDNER, ziel_files[-1]),
                sep=";",
                encoding="utf-8-sig",
                dtype=str,
            )
            merged = pd.merge(start_df, ziel_df, on="Startnummer", suffixes=("_start", "_ziel"))

            def berechne_zeit(row):
                try:
                    t1 = datetime.strptime(row["Startzeit"], "%H:%M:%S.%f")
                    t2 = datetime.strptime(row["Zielzeit"], "%H:%M:%S.%f")
                    delta = t2 - t1
                    if delta < timedelta(0):
                        return pd.Series([None, None])
                    total_seconds = delta.total_seconds()
                    minutes = int(total_seconds // 60)
                    seconds = int(total_seconds % 60)
                    hundredths = int((total_seconds - int(total_seconds)) * 100)
                    formatted = f"{minutes:02}:{seconds:02}:{hundredths:02}"
                    return pd.Series([formatted, total_seconds])
                except Exception:
                    return pd.Series([None, None])

            merged[["Rennzeit", "Rennzeit_Sekunden"]] = merged.apply(berechne_zeit, axis=1)
            merged = merged.dropna(subset=["Rennzeit"])
            merged["km/h"] = (DISTANZ_M / merged["Rennzeit_Sekunden"]) * 3.6
            merged["Kategorie"] = merged.apply(bestimme_kategorie, axis=1)
            return merged

        aktuelles_jahr = datetime.now().year
        junior_jahrgang = aktuelles_jahr - 15
        nachwuchs_jahrgang = aktuelles_jahr - 12

        def bestimme_kategorie(row):
            geschlecht = row.get("Geschlecht", "")
            club = row.get("Clubmitglied", "")
            jahrgang_str = row.get("Jahrgang_start") or row.get("Jahrgang_ziel") or "0"
            try:
                jahrgang = int(jahrgang_str)
            except ValueError:
                jahrgang = 0

            if club != "Ja" and jahrgang < junior_jahrgang:
                return "Gäste Herren" if geschlecht == "Männlich" else "Gäste Damen"

            if geschlecht == "Männlich":
                if jahrgang >= nachwuchs_jahrgang:
                    return "Nachwuchs Knaben"
                if jahrgang >= junior_jahrgang:
                    return "Junioren"
                return "Herren"
            else:
                if jahrgang >= nachwuchs_jahrgang:
                    return "Nachwuchs Mädchen"
                if jahrgang >= junior_jahrgang:
                    return "Juniorinnen"
                return "Damen"

        def zeit_in_hundertstel(zeit_str: str) -> int:
            m, s, h = map(int, zeit_str.split(":"))
            return (m * 60 + s) * 100 + h

        def format_hundertstel(cs: int) -> str:
            minuten = cs // 6000
            sekunden = (cs % 6000) // 100
            hundertstel = cs % 100
            return f"{minuten:02}:{sekunden:02}:{hundertstel:02}"

        def kategorisieren(df):
            kategorien_dfs = {}
            for kat in KATEGORIEN_REIHENFOLGE:
                df_kat = df[df["Kategorie"] == kat].copy()
                if df_kat.empty:
                    continue
                df_kat = df_kat.sort_values(by="Rennzeit_Sekunden").reset_index(drop=True)
                df_kat["Rang"] = df_kat.index + 1
                basis = zeit_in_hundertstel(df_kat["Rennzeit"].iloc[0])
                df_kat["Rückstand"] = df_kat["Rennzeit"].apply(
                    lambda z: format_hundertstel(zeit_in_hundertstel(z) - basis)
                )
                df_kat.rename(
                    columns={
                        "Vorname_ziel": "Vorname",
                        "Nachname_ziel": "Nachname",
                        "Jahrgang_ziel": "Jahrgang",
                        "Wohnort_ziel": "Wohnort",
                    },
                    inplace=True,
                )
                if "km/h" in df_kat.columns:
                    df_kat["km/h"] = df_kat["km/h"].round(2)
                if {"Rennzeit_flach", "Rennzeit_berg"}.issubset(df_kat.columns):
                    df_kat.rename(
                        columns={
                            "Rennzeit_flach": "Zeit Flach",
                            "Rennzeit_berg": "Zeit Berg",
                        },
                        inplace=True,
                    )
                    df_kat = df_kat[
                        [
                            "Rang",
                            "Vorname",
                            "Nachname",
                            "Jahrgang",
                            "Wohnort",
                            "Zeit Flach",
                            "Zeit Berg",
                            "Rennzeit",
                            "Rückstand",
                        ]
                    ]
                else:
                    df_kat = df_kat[
                        [
                            "Rang",
                            "Vorname",
                            "Nachname",
                            "Jahrgang",
                            "Wohnort",
                            "Rennzeit",
                            "Rückstand",
                            "km/h",
                        ]
                    ]
                kategorien_dfs[kat] = df_kat
            return kategorien_dfs

        flach_df = lade_race("Flach")
        berg_df = lade_race("Berg")

        if flach_df is None or berg_df is None:
            messagebox.showerror("Fehler", "Start- oder Zielzeit-Datei(en) fehlen.")
            return

        flach_kats = kategorisieren(flach_df)
        berg_kats = kategorisieren(berg_df)

        def format_seconds(sec):
            m = int(sec // 60)
            s = int(sec % 60)
            h = int((sec - int(sec)) * 100)
            return f"{m:02}:{s:02}:{h:02}"

        gesamt = pd.merge(
            flach_df[
                [
                    "Startnummer",
                    "Vorname_ziel",
                    "Nachname_ziel",
                    "Jahrgang_ziel",
                    "Wohnort_ziel",
                    "Geschlecht",
                    "Clubmitglied",
                    "Rennzeit",
                    "Rennzeit_Sekunden",
                ]
            ],
            berg_df[["Startnummer", "Rennzeit", "Rennzeit_Sekunden"]],
            on="Startnummer",
            suffixes=("_flach", "_berg"),
        )

        gesamt["Rennzeit_Sekunden"] = (
            gesamt["Rennzeit_Sekunden_flach"] + gesamt["Rennzeit_Sekunden_berg"]
        )
        gesamt["Rennzeit"] = gesamt["Rennzeit_Sekunden"].apply(format_seconds)
        gesamt["Kategorie"] = gesamt.apply(bestimme_kategorie, axis=1)

        gesamt_kats = kategorisieren(gesamt)

        def remove_header_border(sheet, startrow, df):
            header_row = startrow + 2
            for col in range(1, df.shape[1] + 1):
                sheet.cell(row=header_row, column=col).border = Border()

        with pd.ExcelWriter(AUSGABE_DATEI, engine="openpyxl") as writer:
            startrow = 0
            for kat, df_kat in flach_kats.items():
                df_kat.to_excel(writer, sheet_name="Flachrennen", startrow=startrow + 1, index=False)
                sheet = writer.sheets["Flachrennen"]
                remove_header_border(sheet, startrow, df_kat)
                sheet.cell(row=startrow + 1, column=1, value=kat)
                startrow += len(df_kat) + 3

            startrow = 0
            for kat, df_kat in berg_kats.items():
                df_kat.to_excel(writer, sheet_name="Bergrennen", startrow=startrow + 1, index=False)
                sheet = writer.sheets["Bergrennen"]
                remove_header_border(sheet, startrow, df_kat)
                sheet.cell(row=startrow + 1, column=1, value=kat)
                startrow += len(df_kat) + 3

            startrow = 0
            for kat, df_kat in gesamt_kats.items():
                df_kat.to_excel(
                    writer, sheet_name="Gesamtwertung", startrow=startrow + 1, index=False
                )
                sheet = writer.sheets["Gesamtwertung"]
                remove_header_border(sheet, startrow, df_kat)
                sheet.cell(row=startrow + 1, column=1, value=kat)
                startrow += len(df_kat) + 3

            for name in ["Flachrennen", "Bergrennen", "Gesamtwertung"]:
                ws = writer.sheets[name]
                ws.sheet_view.view = "pageLayout"
                ws.sheet_view.showGridLines = False
                ws.print_options.gridLines = False
                ws.page_margins = PageMargins(
                    left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.3, footer=0.3
                )
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left")
                        cell.border = Border()
                h = ws.oddHeader
                h.left.text = name
                h.left.size = 20

        text_output = ""
        def append_text(title, kat_dict):
            nonlocal text_output
            text_output += f"{title}\n"
            for kat, df_kat in kat_dict.items():
                text_output += f"{kat}\n"
                text_output += df_kat.to_string(index=False)
                text_output += "\n\n"

        append_text("Flachrennen", flach_kats)
        append_text("Bergrennen", berg_kats)
        append_text("Gesamtwertung", gesamt_kats)

        self.anzeige.delete("1.0", "end")
        self.anzeige.insert("1.0", text_output.strip())
